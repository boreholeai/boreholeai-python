"""Fan-out + poll + download orchestration for SDK batch processing.

Reads/writes the manifest, calls APIClientAsync for HTTP. Three phases
run sequentially, but within each phase all files are processed
concurrently:

   SUBMIT    bounded by Semaphore(concurrency)
   POLL      unbounded — polling is idle wait, not real load
   DOWNLOAD  bounded by Semaphore(concurrency)

State is checkpointed to the manifest after every transition, so a
Ctrl-C or network drop can be resumed by re-invoking with the same
output_dir.
"""

from __future__ import annotations

import asyncio
import logging
import random
import shutil
import time
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Callable, Coroutine, Optional

import httpx

from boreholeai import _manifest
from boreholeai._api import APIClientAsync
from boreholeai._manifest import (
    STATUS_COMPLETED,
    STATUS_FAILED,
    STATUS_PROCESSING,
    STATUS_QUEUED,
    STATUS_SUBMITTED,
    STATUS_SUBMIT_FAILED,
    Manifest,
)
from boreholeai.exceptions import (
    AuthenticationError,
    BoreholeAIError,
    InsufficientCreditsError,
    RateLimitError,
    ServerError,
)

logger = logging.getLogger(__name__)

# Client-side ceiling only — the effective in-flight count is
# min(this, the API key's server-side max_concurrent_jobs from /v1/me),
# so the admin-page cap governs whenever it is <= this value.
_DEFAULT_CONCURRENCY = 20

_POLL_INITIAL_INTERVAL = 2.0
_POLL_MAX_INTERVAL = 10.0
_POLL_BACKOFF_FACTOR = 1.5

# Submit retries — safety net for genuine transient errors (5xx, brief
# network blips). 5 attempts with exponential backoff.
_SUBMIT_MAX_RETRIES = 5
_SUBMIT_RETRY_BASE = 2.0
_SUBMIT_RETRY_MAX = 60.0

# 429 on submit is backpressure, not an error: with cap-aware pacing the
# client believes it holds a slot, so a 429 means the server's in-flight
# count disagrees (stale rows, a parallel session, a web upload). Slots
# free on job-completion timescales (minutes), so the file waits its turn
# with jittered backoff and NO attempt budget. The only bound is batch
# liveness: if nothing in the whole run has reserved a slot or reached a
# terminal state for this long while submits are still refused, the cap
# is provably stuck and waiting files fail with a pointed message.
_RATE_LIMIT_RETRY_BASE = 5.0    # matches the server's Retry-After hint
_RATE_LIMIT_RETRY_MAX = 60.0
_RATE_LIMIT_STALL_BUDGET = 30 * 60.0

# Stuck-job guard. A live job's poll responses change constantly (status
# transitions, page/subgraph progress stamps, and — on servers that expose
# updated_at — a once-a-minute worker liveness heartbeat); a job whose
# worker died mid-flight answers "processing" with frozen fields forever.
# If nothing observable changes for the budget, fail the file client-side
# and release its slot. 35 min is deliberately ABOVE the server's stale-job
# self-heal threshold (10 min of heartbeat silence + a <=5 min sweep), so
# the server gets first crack: a rescued job resumes invisibly (the requeue
# resets this clock) and the client guard only fires when the server
# couldn't help. Queued jobs get a longer leash: waiting behind a deep
# queue is legitimate stillness.
_STUCK_PROCESSING_BUDGET = 35 * 60.0
_STUCK_QUEUED_BUDGET = 120 * 60.0

# Download retries — bounded, for transient network errors while fetching
# results. Exhausted retries mark the file failed for this run; a re-run
# re-attempts the download without resubmitting the job.
_DOWNLOAD_MAX_RETRIES = 3
_DOWNLOAD_RETRY_BASE = 2.0
_DOWNLOAD_RETRY_MAX = 30.0

# Poll error taxonomy. Permanent HTTP verdicts fail the file immediately —
# retrying cannot change the server's answer (revoked key, no credits,
# job unknown). Everything else retries with backoff but gives up after
# this many CONSECUTIVE errors (~5 min at the max interval); a successful
# poll resets the streak, so a long server-side queue still waits
# indefinitely.
_POLL_PERMANENT_HTTP = (401, 402, 404)
_POLL_MAX_ERROR_STREAK = 30

# Statuses the server can legitimately report on a poll. Anything else is
# treated as a server fault (counts toward the error streak) and is never
# written into the manifest, so a resumed run can't be stranded on a
# status the state machine doesn't know.
_KNOWN_POLL_STATUSES = (
    STATUS_SUBMITTED, STATUS_QUEUED, STATUS_PROCESSING,
    STATUS_COMPLETED, STATUS_FAILED,
)

_WORKDIR_NAME = ".boreholeai_workdir"
_UNSAFE_WORKDIR_CHARS = '<>:"/\\|?*' + "".join(
    chr(c) for c in range(0x00, 0x20)
)
_WORKDIR_TRANS = str.maketrans({c: "_" for c in _UNSAFE_WORKDIR_CHARS})


def _job_workdir_candidates(
    workdir: Path, filename: str, job_id: str,
) -> tuple[Path, Path]:
    """Return the labelled and legacy cache paths for one job."""
    labelled_component = f"{filename} {job_id}".translate(_WORKDIR_TRANS)
    legacy_component = str(job_id).translate(_WORKDIR_TRANS)
    return workdir / labelled_component, workdir / legacy_component


def _job_workdir(workdir: Path, filename: str, job_id: str) -> Path:
    """Return the per-job cache directory.

    New SDK runs use ``<filename> <job_id>`` so a person inspecting the
    cache can identify each job without opening the manifest. If an older
    job-id-only directory already exists, keep reading/writing it so an SDK
    upgrade never breaks an in-progress resume.
    """
    labelled, legacy = _job_workdir_candidates(workdir, filename, job_id)
    if not labelled.exists() and legacy.exists():
        return legacy
    return labelled


def _remove_job_workdirs(workdir: Path, filename: str, job_id: str) -> None:
    """Remove all client-side cache layouts belonging to an old job."""
    for path in _job_workdir_candidates(workdir, filename, job_id):
        if path.is_dir():
            shutil.rmtree(path)
        elif path.exists():
            path.unlink()


@dataclass
class BatchResult:
    """Outcome of `run_batch`. Filenames are input file basenames."""

    successes: list[str] = field(default_factory=list)
    failures: dict[str, str] = field(default_factory=dict)
    job_ids: dict[str, str] = field(default_factory=dict)
    workdir: Optional[Path] = None
    manifest: Optional[Manifest] = None


class _BatchPulse:
    """Batch-wide liveness clock for 429 backpressure decisions.

    Beats when server-side capacity provably moved: a submit reserved a
    slot, or a poll saw a job reach a terminal state (which frees one).
    Files waiting out a 429 consult `stalled_for()` to distinguish "cap
    full but flowing" (keep waiting) from "cap stuck" (give up).
    """

    def __init__(self) -> None:
        self.last_progress = time.monotonic()

    def beat(self) -> None:
        self.last_progress = time.monotonic()

    def stalled_for(self) -> float:
        return time.monotonic() - self.last_progress


async def run_batch(
    client: APIClientAsync,
    files: list[Path],
    output_dir: Path,
    *,
    concurrency: int = _DEFAULT_CONCURRENCY,
    on_progress: Optional[Callable[[Manifest], None]] = None,
) -> BatchResult:
    """Submit, poll, and download N files concurrently.

    Persists state to `output_dir/.boreholeai_manifest.json` so an
    interrupted run can resume by re-invoking with the same arguments.
    Does NOT call merge — caller decides when to merge across the
    workdir.

    Caller is expected to have already passed an effective `concurrency`
    that respects the server's per-user cap (see
    `resolve_effective_concurrency`). `run_batch` itself does not call
    `/v1/me`.

    `on_progress` (optional): called with the current Manifest after every
    state change. Used by `client.py` to render a per-file progress display.
    Must not raise; exceptions are swallowed to avoid breaking the batch.
    """
    files = list(files)
    if not files:
        raise ValueError("run_batch requires at least one file")

    output_dir = Path(output_dir).resolve()
    output_dir.mkdir(parents=True, exist_ok=True)

    workdir = output_dir / _WORKDIR_NAME
    workdir.mkdir(parents=True, exist_ok=True)

    input_root = files[0].parent.resolve()
    manifest = _manifest.load_or_init(
        output_dir, input_root=input_root, concurrency=concurrency, files=files,
    )
    save_lock = asyncio.Lock()
    cap_sem = asyncio.Semaphore(concurrency)         # gates server-side cap
    download_sem = asyncio.Semaphore(concurrency)    # bounds download bandwidth

    files_by_name = {f.name: f for f in files}

    # Initial render so users see all files in their starting state
    _emit_progress(on_progress, manifest)

    # Per-file lifecycle: submit → poll under cap_sem (matches server cap),
    # then download under download_sem (separate — once a job hits a terminal
    # status it no longer consumes the server cap, so we don't have to hold
    # cap_sem through download). Result: file 1's download runs in parallel
    # with file 2's submit+poll, instead of serializing.
    pulse = _BatchPulse()
    await asyncio.gather(*(
        _process_one_safe(
            name, files_by_name[name], client,
            cap_sem, download_sem,
            manifest, save_lock, output_dir, workdir, pulse, on_progress,
        )
        for name in files_by_name
    ))

    return _build_result(manifest, files_by_name, workdir)


async def _process_one_safe(
    name: str, file_path: Path, client: APIClientAsync,
    cap_sem: asyncio.Semaphore, download_sem: asyncio.Semaphore,
    manifest: Manifest, save_lock: asyncio.Lock,
    output_dir: Path, workdir: Path, pulse: _BatchPulse,
    on_progress: Optional[Callable[[Manifest], None]] = None,
) -> None:
    """Last-resort firewall around one file's lifecycle.

    Every anticipated error class is handled inside `_process_one`; this
    net exists so that an unanticipated one (server contract drift, local
    I/O surprise) fails this one file instead of cancelling the whole
    gather and killing the other N-1 files mid-run.
    """
    try:
        await _process_one(
            name, file_path, client, cap_sem, download_sem,
            manifest, save_lock, output_dir, workdir, pulse, on_progress,
        )
    except Exception as exc:
        logger.error("unexpected error on %s: %r", name, exc)
        async with save_lock:
            e = manifest.jobs[name]
            e.status = STATUS_FAILED
            e.error = f"unexpected error: {exc!r}"
            _manifest.save(manifest, output_dir)
        _emit_progress(on_progress, manifest)


async def _process_one(
    name: str, file_path: Path, client: APIClientAsync,
    cap_sem: asyncio.Semaphore, download_sem: asyncio.Semaphore,
    manifest: Manifest, save_lock: asyncio.Lock,
    output_dir: Path, workdir: Path, pulse: _BatchPulse,
    on_progress: Optional[Callable[[Manifest], None]] = None,
) -> None:
    """Per-file lifecycle: submit → poll → download.

    `cap_sem` is held across submit + poll only, so file B can't even
    attempt to submit until file A reaches a terminal status on the server.
    Download runs after `cap_sem` is released, under a separate semaphore
    that bounds bandwidth without blocking the next file's submission.
    """
    # Submit + poll under cap_sem (server-side concurrency)
    async with cap_sem:
        if manifest.needs_submit(name):
            await _submit_one(
                name, file_path, client, manifest, save_lock, output_dir,
                pulse, on_progress,
            )
        # If submit succeeded, status is now SUBMITTED → needs_poll is True.
        # If submit failed, no job_id was set → needs_poll is False, skip.
        if manifest.needs_poll(name):
            await _poll_one(
                name, client, manifest, save_lock, output_dir, pulse, on_progress,
            )

    # Download under download_sem (no server cap consumed; just bandwidth)
    if manifest.needs_download(name):
        async with download_sem:
            await _download_one(
                name, client, manifest, save_lock, output_dir, workdir, on_progress,
            )




# -------------------------------------------
# Internal Helper Functions
# -------------------------------------------

async def _submit_one(
    name: str, file_path: Path, client: APIClientAsync,
    manifest: Manifest,
    save_lock: asyncio.Lock, output_dir: Path,
    pulse: _BatchPulse,
    on_progress: Optional[Callable[[Manifest], None]] = None,
) -> None:
    """Submit one file, waiting out 429 backpressure and retrying
    transient errors with exponential backoff.

    Caller is expected to hold the per-file semaphore (`_submit_and_poll`
    wraps this in `async with sem:`) so the slot is reserved for the
    entire submit + poll lifecycle, matching the server's cap semantics.
    """
    delay = _SUBMIT_RETRY_BASE
    rl_delay = _RATE_LIMIT_RETRY_BASE
    transient_attempts = 0
    while True:
        try:
            data = await client.create_job([file_path])
            job_id = data.get("job_id") if isinstance(data, dict) else None
            if not job_id:
                # Server accepted the POST but broke its contract — treat
                # like a 5xx so the existing transient retry applies.
                raise ServerError("Server accepted the upload but returned no job id")
            pulse.beat()  # a slot was reserved — the cap is flowing
            async with save_lock:
                e = manifest.jobs[name]
                old_job_id = e.job_id if e.reprocess else None
                e.job_id = job_id
                e.num_pages = data.get("num_pages")
                e.status = STATUS_SUBMITTED
                e.submitted_at = _manifest._now()
                e.local_data = client.local_data
                e.error = None
                # A (re)submitted entry is a new server job — any download
                # state from a previous job_id no longer applies, and a
                # user-set reprocess flag is now consumed. Without the
                # downloaded reset, a manually re-queued entry would never
                # have its new results downloaded.
                e.downloaded = False
                e.purged = False
                e.reprocess = False
                e.warning = None
                e.completed_at = None
                e.pages_done = 0
                e.current_page = 1
                e.pages_total = e.num_pages or 1
                e.completed_subgraphs = []
                _manifest.save(manifest, output_dir)
            if old_job_id and old_job_id != job_id:
                try:
                    _remove_job_workdirs(
                        output_dir / _WORKDIR_NAME, name, old_job_id,
                    )
                    logger.info(
                        "removed old client cache for %s → job %s",
                        name, old_job_id[:8],
                    )
                except OSError as exc:
                    # The replacement job is already accepted and persisted.
                    # Keep processing it; a stale cache folder is isolated by
                    # job ID and can never participate in the active merge.
                    logger.warning(
                        "could not remove old client cache for %s (job %s): %s",
                        name, old_job_id[:8], exc,
                    )
            _emit_progress(on_progress, manifest)
            logger.info("submitted %s → job %s", name, job_id[:8])
            return
        except RateLimitError as exc:
            # Backpressure, not an error: the server's in-flight count is
            # at the cap (stale rows, a parallel session, a web upload).
            # Slots free on job-completion timescales, so wait — jittered
            # so concurrent waiters don't retry in synchronised waves —
            # for as long as the batch shows life. No attempt budget.
            if pulse.stalled_for() > _RATE_LIMIT_STALL_BUDGET:
                await _mark_submit_failed(
                    name,
                    "concurrency cap appears stuck — no job in this run "
                    f"reserved a slot or completed for {int(_RATE_LIMIT_STALL_BUDGET // 60)} "
                    "minutes while submits were rate-limited. Check the admin "
                    "jobs list for stale queued/processing jobs.",
                    manifest, save_lock, output_dir, on_progress,
                )
                return
            wait = rl_delay * (0.5 + random.random())
            logger.info(
                "cap full — %s waiting %.1fs for a slot (%s)", name, wait, exc,
            )
            await asyncio.sleep(wait)
            rl_delay = min(rl_delay * 2, _RATE_LIMIT_RETRY_MAX)
        except (ServerError, httpx.TransportError) as exc:
            transient_attempts += 1
            if transient_attempts >= _SUBMIT_MAX_RETRIES:
                await _mark_submit_failed(name, str(exc), manifest, save_lock, output_dir, on_progress)
                return
            logger.warning(
                "transient submit error on %s (%s) — retry %d/%d in %.1fs",
                name, exc, transient_attempts, _SUBMIT_MAX_RETRIES, delay,
            )
            await asyncio.sleep(delay)
            delay = min(delay * 2, _SUBMIT_RETRY_MAX)
        except OSError as exc:
            # Local read failure — the file was deleted/renamed/made
            # unreadable after collection (e.g. mid-run folder edit). Not
            # transient: retrying re-reads the same missing file. Fail this
            # file only; if it reappears, a later re-run picks it up.
            await _mark_submit_failed(
                name, f"source file missing or unreadable: {exc}",
                manifest, save_lock, output_dir, on_progress,
            )
            return
        except (AuthenticationError, InsufficientCreditsError) as exc:
            # Fatal — no point retrying this file or any other (caller will see)
            await _mark_submit_failed(name, str(exc), manifest, save_lock, output_dir, on_progress)
            return
        except BoreholeAIError as exc:
            await _mark_submit_failed(name, str(exc), manifest, save_lock, output_dir, on_progress)
            return


async def _mark_submit_failed(
    name: str, error: str, manifest: Manifest,
    save_lock: asyncio.Lock, output_dir: Path,
    on_progress: Optional[Callable[[Manifest], None]] = None,
) -> None:
    async with save_lock:
        e = manifest.jobs[name]
        e.status = STATUS_SUBMIT_FAILED
        e.error = error
        _manifest.save(manifest, output_dir)
    _emit_progress(on_progress, manifest)
    logger.error("submit failed for %s: %s", name, error)


async def _poll_one(
    name: str, client: APIClientAsync, manifest: Manifest,
    save_lock: asyncio.Lock, output_dir: Path,
    pulse: Optional[_BatchPulse] = None,
    on_progress: Optional[Callable[[Manifest], None]] = None,
) -> None:
    """Poll one job until it reaches a terminal state.

    Transient errors (timeouts, 429/5xx, unparseable bodies, unknown
    statuses) retry with backoff, bounded to `_POLL_MAX_ERROR_STREAK`
    consecutive failures. Permanent HTTP verdicts (401/402/404) fail the
    file immediately so the concurrency slot is released — retrying a
    revoked key or a deleted job forever would hang the batch silently.
    """
    job_id = manifest.jobs[name].job_id
    if job_id is None:
        return

    interval = _POLL_INITIAL_INTERVAL
    error_streak = 0
    last_observed: Optional[tuple] = None
    last_change = time.monotonic()
    while True:
        try:
            data = await client.get_job(job_id)
        except (BoreholeAIError, httpx.TransportError) as exc:
            if getattr(exc, "status_code", None) in _POLL_PERMANENT_HTTP:
                await _mark_poll_failed(
                    name, f"poll rejected (HTTP {exc.status_code}): {exc}",
                    manifest, save_lock, output_dir, on_progress,
                )
                return
            error_streak += 1
            if error_streak >= _POLL_MAX_ERROR_STREAK:
                await _mark_poll_failed(
                    name, f"poll gave up after {error_streak} consecutive errors: {exc}",
                    manifest, save_lock, output_dir, on_progress,
                )
                return
            logger.warning(
                "poll error on %s (%s) — retry %d/%d in %.1fs",
                name, exc, error_streak, _POLL_MAX_ERROR_STREAK, interval,
            )
            await asyncio.sleep(interval)
            interval = min(interval * _POLL_BACKOFF_FACTOR, _POLL_MAX_INTERVAL)
            continue

        status = data.get("status") if isinstance(data, dict) else None
        if status not in _KNOWN_POLL_STATUSES:
            # Contract drift or a mangled body — don't write it into the
            # manifest (a resumed run couldn't act on it); treat like any
            # other transient server fault.
            error_streak += 1
            if error_streak >= _POLL_MAX_ERROR_STREAK:
                await _mark_poll_failed(
                    name, f"server kept returning unrecognised status {status!r}",
                    manifest, save_lock, output_dir, on_progress,
                )
                return
            logger.warning(
                "unrecognised status %r on %s — retry %d/%d in %.1fs",
                status, name, error_streak, _POLL_MAX_ERROR_STREAK, interval,
            )
            await asyncio.sleep(interval)
            interval = min(interval * _POLL_BACKOFF_FACTOR, _POLL_MAX_INTERVAL)
            continue

        error_streak = 0
        progress = data.get("progress") or {}
        if not isinstance(progress, dict):
            progress = {}
        async with save_lock:
            e = manifest.jobs[name]
            e.status = status
            # Live page + subgraph state for the renderer (drives the
            # progress bar via _progress.compute_progress). Progress is
            # cosmetic — junk values are ignored, never fatal.
            page = _as_int(progress.get("page"))
            pages_total = _as_int(progress.get("pages_total"))
            completed = progress.get("completed_subgraphs")
            if page is not None:
                e.current_page = page
            if pages_total:
                e.pages_total = pages_total
                if e.num_pages is None:
                    e.num_pages = pages_total
            if isinstance(completed, list):
                e.completed_subgraphs = [str(s) for s in completed]
            # Legacy pages_done — kept up to date for older renderers.
            pages_done = _as_int(progress.get("pages_done"))
            if pages_done is not None:
                e.pages_done = pages_done
            if status == STATUS_FAILED:
                e.error = data.get("error_message", "unknown error")
                e.completed_at = _manifest._now()
            elif status == STATUS_COMPLETED:
                e.completed_at = _manifest._now()
                if e.num_pages is None:
                    e.num_pages = _as_int(data.get("num_pages"))
                if e.num_pages:
                    e.pages_done = e.num_pages
            observed = (
                status, e.current_page, e.pages_total,
                tuple(e.completed_subgraphs), e.pages_done,
                # Server liveness heartbeat (newer servers stamp this every
                # minute while a worker is alive). Absent on older servers,
                # where the guard falls back to progress changes alone.
                data.get("updated_at"),
            )
            _manifest.save(manifest, output_dir)
        _emit_progress(on_progress, manifest)

        if status in (STATUS_COMPLETED, STATUS_FAILED):
            if pulse is not None:
                pulse.beat()  # terminal state frees a server-side slot
            return

        # Stuck-job guard: the poll loop is the timer — remember what we
        # last saw and when it last changed. A dead worker's job answers
        # "processing" with frozen progress forever; a live one changes
        # within minutes.
        if observed != last_observed:
            last_observed = observed
            last_change = time.monotonic()
        else:
            budget = (
                _STUCK_PROCESSING_BUDGET if status == STATUS_PROCESSING
                else _STUCK_QUEUED_BUDGET
            )
            if time.monotonic() - last_change > budget:
                await _mark_poll_failed(
                    name,
                    f"job appears stuck — no progress for "
                    f"{int(budget // 60)} minutes in status {status!r}. The "
                    "worker may have died mid-job; requeue it from the admin "
                    "jobs list or restart the worker server.",
                    manifest, save_lock, output_dir, on_progress,
                )
                return

        await asyncio.sleep(interval)
        interval = min(interval * _POLL_BACKOFF_FACTOR, _POLL_MAX_INTERVAL)


async def _mark_poll_failed(
    name: str, error: str, manifest: Manifest,
    save_lock: asyncio.Lock, output_dir: Path,
    on_progress: Optional[Callable[[Manifest], None]] = None,
) -> None:
    """Record a poll-level permanent failure. Status FAILED means the
    re-run loop treats the file as re-submittable once the underlying
    cause (key, credits, deleted job) is fixed."""
    async with save_lock:
        e = manifest.jobs[name]
        e.status = STATUS_FAILED
        e.error = error
        e.completed_at = _manifest._now()
        _manifest.save(manifest, output_dir)
    _emit_progress(on_progress, manifest)
    logger.error("poll failed for %s: %s", name, error)


def _as_int(value: Any) -> Optional[int]:
    """Best-effort int conversion; None for junk instead of ValueError."""
    try:
        return int(value)
    except (TypeError, ValueError):
        return None


async def _download_one(
    name: str, client: APIClientAsync,
    manifest: Manifest, save_lock: asyncio.Lock,
    output_dir: Path, workdir: Path,
    on_progress: Optional[Callable[[Manifest], None]] = None,
) -> None:
    """Fetch signed URLs, download files into
    ``workdir/<filename> <job_id>/``, and optionally call DELETE for the
    enterprise purge-on-download flow.

    Caller is expected to hold the per-file download semaphore
    (`_process_one` wraps this in `async with download_sem:`).
    """
    job_id = manifest.jobs[name].job_id
    if job_id is None:
        return

    job_workdir = _job_workdir(workdir, name, job_id)
    job_workdir.mkdir(parents=True, exist_ok=True)

    try:
        results = await _retry_transport(
            lambda: client.get_results(job_id), "results fetch", name,
        )
    except (BoreholeAIError, httpx.TransportError) as exc:
        async with save_lock:
            manifest.jobs[name].error = f"results fetch failed: {exc}"
            _manifest.save(manifest, output_dir)
        _emit_progress(on_progress, manifest)
        logger.error("results fetch failed for %s: %s", name, exc)
        return

    files_to_dl = results.get("files") if isinstance(results, dict) else None
    if not isinstance(files_to_dl, list) or not files_to_dl:
        # A completed job always has result files. A missing/malformed/empty
        # list means the server broke its contract — fail the file (leaving
        # downloaded=False so a re-run re-attempts) rather than "succeeding"
        # with an empty job dir that silently drops this borehole from the
        # merge.
        async with save_lock:
            manifest.jobs[name].error = (
                f"results fetch failed: invalid or empty file list ({files_to_dl!r})"
            )
            _manifest.save(manifest, output_dir)
        _emit_progress(on_progress, manifest)
        logger.error("invalid results file list for %s: %r", name, files_to_dl)
        return
    for f in files_to_dl:
        try:
            entry = f if isinstance(f, dict) else {}
            # Flatten to a basename: a server-side bug must not be able to
            # write outside the job workdir via ../ or an absolute path.
            filename = Path(str(entry.get("filename") or "")).name
            url = entry.get("url")
            if not filename or not url:
                raise ValueError(f"malformed file entry in results: {f!r}")
            content = await _retry_transport(
                lambda: client.download_file(url), "download", name,
            )
            (job_workdir / filename).write_bytes(content)
        except Exception as exc:
            async with save_lock:
                manifest.jobs[name].error = f"download failed: {exc}"
                _manifest.save(manifest, output_dir)
            _emit_progress(on_progress, manifest)
            logger.error("download failed for %s: %s", name, exc)
            return

    # Job-level success can hide page-level extraction failures (a page
    # that fails validation is skipped; the job still completes). The
    # job's own Processing Info sheet records them — surface that.
    warning = _page_failure_warning(job_workdir)

    async with save_lock:
        e = manifest.jobs[name]
        e.downloaded = True
        e.warning = warning
        _manifest.save(manifest, output_dir)
    _emit_progress(on_progress, manifest)
    logger.info("downloaded %d file(s) for %s", len(files_to_dl), name)
    if warning:
        logger.warning("%s: %s", name, warning)

    if results.get("purge_on_download"):
        try:
            purge = await client.delete_job(job_id)
            async with save_lock:
                manifest.jobs[name].purged = True
                _manifest.save(manifest, output_dir)
            _emit_progress(on_progress, manifest)
            logger.info(
                "purged %d server file(s) for %s",
                purge.get("files_deleted", 0), name,
            )
        except BoreholeAIError as exc:
            logger.warning("purge failed for %s: %s", name, exc)


def _page_failure_warning(job_workdir: Path) -> Optional[str]:
    """Describe page-level extraction failures from the job's own
    Processing Info sheet, or None if every page digitised.

    Never raises: a completed, downloaded job must not fail because this
    diagnostic couldn't be read (missing sheet, odd values, corrupt file
    are all treated as "nothing to report").
    """
    try:
        from openpyxl import load_workbook

        from boreholeai._merge._processing_info import read_processing_info

        gp = next(iter(sorted(job_workdir.glob("Borehole_ground_profile*.xlsx"))), None)
        if gp is None:
            return None
        wb = load_workbook(gp, read_only=True, data_only=True)
        try:
            if "Processing Info" not in wb.sheetnames:
                return None
            info = read_processing_info(wb["Processing Info"])
        finally:
            wb.close()
        failed = int(str(info.get("Boreholes Failed", "0")).strip())
        if failed <= 0:
            return None
        total = str(info.get("Total Boreholes", "?")).strip() or "?"
        names = [k.strip() for k in info if k.startswith("  ")]
        listing = f": {', '.join(names)}" if names else ""
        return (
            f"{failed} of {total} borehole page(s) failed extraction"
            f"{listing} — see the Processing Info sheet in the Excel"
        )
    except Exception:
        return None


async def _retry_transport(
    fn: Callable[[], Coroutine[Any, Any, Any]], what: str, name: str,
) -> Any:
    """Await `fn()`, retrying transient network errors with backoff.

    Only retries httpx.TransportError (timeouts, dropped connections);
    HTTP-level and SDK errors propagate to the caller unchanged.
    """
    delay = _DOWNLOAD_RETRY_BASE
    for attempt in range(1, _DOWNLOAD_MAX_RETRIES + 1):
        try:
            return await fn()
        except httpx.TransportError as exc:
            if attempt >= _DOWNLOAD_MAX_RETRIES:
                raise
            logger.warning(
                "transient %s error on %s (%s) — retry %d/%d in %.1fs",
                what, name, exc, attempt, _DOWNLOAD_MAX_RETRIES, delay,
            )
            await asyncio.sleep(delay)
            delay = min(delay * 2, _DOWNLOAD_RETRY_MAX)


def _emit_progress(
    cb: Optional[Callable[[Manifest], None]], manifest: Manifest,
) -> None:
    """Fire the progress callback if set, swallowing any exceptions so the
    batch never fails because the renderer crashed."""
    if cb is None:
        return
    try:
        cb(manifest)
    except Exception:
        logger.debug("progress callback raised", exc_info=True)


async def resolve_effective_concurrency(
    client: APIClientAsync, user_concurrency: int,
) -> tuple[int, Optional[int]]:
    """Decide the SDK's submit-semaphore size, capped by the server.

    Calls GET /v1/me. Returns (effective_concurrency, server_cap_or_None).
    On 4xx/5xx or network errors, falls back to `user_concurrency`,
    returning (user_concurrency, None) — the safety-net 429 retries
    in `_submit_one` will catch any cap-overshoot.

    Raises `BoreholeAIError` if the server reports cap=0 (admin lockout).
    """
    try:
        info = await client.get_me()
    except Exception as exc:
        logger.warning(
            "could not fetch server concurrency cap (%s) — using "
            "concurrency=%d; cap-rejected requests will retry on 429",
            exc, user_concurrency,
        )
        return user_concurrency, None

    server_cap = info.get("max_concurrent_jobs")
    if server_cap == 0:
        raise BoreholeAIError(
            "your account has no concurrency budget "
            "(max_concurrent_jobs = 0); contact support"
        )
    if not isinstance(server_cap, int) or server_cap < 1:
        logger.warning(
            "server returned unexpected max_concurrent_jobs=%r — "
            "using concurrency=%d", server_cap, user_concurrency,
        )
        return user_concurrency, None

    return min(user_concurrency, server_cap), server_cap


def _build_result(
    manifest: Manifest, files_by_name: dict[str, Path], workdir: Path,
) -> BatchResult:
    res = BatchResult(workdir=workdir, manifest=manifest)
    for name in files_by_name:
        e = manifest.jobs.get(name)
        if e is None:
            continue
        if e.job_id:
            res.job_ids[name] = e.job_id
        if e.status == STATUS_COMPLETED and e.downloaded:
            res.successes.append(name)
        elif e.status in (STATUS_FAILED, STATUS_SUBMIT_FAILED):
            res.failures[name] = e.error or "unknown"
        else:
            res.failures[name] = f"incomplete: status={e.status}"
    return res
