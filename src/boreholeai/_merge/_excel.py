"""Excel data-sheet merging with backend-matching styles.

Mirrors `mergeExcelBuffers` and `applyStyles` in
`frontend/src/app/api/jobs/download/route.ts`.
"""

from __future__ import annotations

import io
import logging
from pathlib import Path
from typing import Iterable, Mapping

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from boreholeai._merge._processing_info import (
    build_merged_processing_info,
    read_processing_info,
)

logger = logging.getLogger(__name__)

# ARGB hex colors — match the backend openpyxl output and the TS constants.
_HEADER_FILL = PatternFill(fill_type="solid", fgColor="FF4472C4")
_HEADER_FONT = Font(bold=True, color="FFFFFFFF", size=11)
_HEADER_SIDE = Side(style="medium", color="FF4472C4")
_HEADER_BORDER = Border(
    top=_HEADER_SIDE, bottom=_HEADER_SIDE, left=_HEADER_SIDE, right=_HEADER_SIDE,
)
_DATA_SIDE = Side(style="thin", color="FFB8CCE4")
_DATA_BORDER = Border(
    top=_DATA_SIDE, bottom=_DATA_SIDE, left=_DATA_SIDE, right=_DATA_SIDE,
)
_ALT_ROW_FILL = PatternFill(fill_type="solid", fgColor="FFF8F9FA")
_FLAGGED_FILL = PatternFill(fill_type="solid", fgColor="FFFFEB9C")
_FLAGGED_FONT = Font(color="FF9C6500")
_FAILED_HEADER_FONT = Font(bold=True, color="FFCC0000", size=11)
_FAILED_NAME_FILL = PatternFill(fill_type="solid", fgColor="FFFFFF00")

_LEFT_ALIGN_COLS = frozenset({
    "Source_File", "Metric", "Value", "material", "material_type", "parent_rock",
    "material_description", "geology_type", "consistency_density_strength",
    "consistency_density", "project_name", "borehole_location",
    "extraction_notes", "grid_datum", "test_results", "other_observations",
    "raw_observations", "SPT_raw", "GWL_date", "Cu_method", "reason_for_review",
})

_WRAP_TEXT_COLS = frozenset({
    "material", "material_description", "reason_for_review",
})


def merge_excel_files(
    paths: Iterable[Path],
    drop_columns: frozenset[str] = frozenset(),
    source_files: Mapping[Path, str] | None = None,
    macro_template: Path | None = None,
) -> bytes:
    """Merge multiple Excel files into one workbook, returned as bytes.

    Sheet rules:
      • "Processing Info" is aggregated via `build_merged_processing_info`.
      • Every other sheet is merged by column name, allowing old and new
        schemas to coexist without shifting values into the wrong columns.
      • `Source_File` is kept first and backfilled from `source_files` for
        legacy workbooks that predate that backend field.
      • Columns whose header is in `drop_columns` are omitted (matched per
        source sheet, so per-file column positions may differ).
      • Styles are applied to every sheet at the end.
      • When `macro_template` is given, that .xlsm — whose Processing Info
        sheet carries the "Regenerate derived tabs" VBA button — is used as
        the base workbook and its VBA project is preserved, so the returned
        bytes are a macro-enabled workbook and must be saved as `.xlsm`.
    """
    paths = list(paths)
    if not paths:
        raise ValueError("merge_excel_files requires at least one path")
    resolved_source_files = {
        Path(path).resolve(): stem
        for path, stem in (source_files or {}).items()
    }

    if macro_template is not None:
        merged = load_workbook(macro_template, keep_vba=True)
    else:
        merged = Workbook()
        # Workbook starts with one default sheet — remove it.
        merged.remove(merged.active)

    # Pass 1: collect Processing Info sheets and build aggregated one.
    processing_infos: list[dict[str, str]] = []
    for p in paths:
        wb = load_workbook(p, data_only=True)
        try:
            if "Processing Info" in wb.sheetnames:
                processing_infos.append(read_processing_info(wb["Processing Info"]))
        finally:
            wb.close()

    if processing_infos:
        # The macro template already carries this sheet (with the VBA
        # button anchored to it) — write into it rather than duplicating.
        if "Processing Info" in merged.sheetnames:
            dest_pi = merged["Processing Info"]
        else:
            dest_pi = merged.create_sheet("Processing Info")
        build_merged_processing_info(processing_infos, dest_pi)

    # Pass 2: collect data sheets as header-addressed records. This prevents
    # positional corruption when old and new workbook schemas are mixed.
    sheet_headers: dict[str, list[str]] = {}
    sheet_rows: dict[str, list[dict[str, object]]] = {}
    for p in paths:
        source_file = resolved_source_files.get(Path(p).resolve())
        wb = load_workbook(p, data_only=True)
        try:
            for name in wb.sheetnames:
                if name == "Processing Info":
                    continue
                src = wb[name]
                rows = src.iter_rows(values_only=True)
                header_row = next(rows, None)
                if header_row is None:
                    continue

                indexed_headers = [
                    (index, str(header))
                    for index, header in enumerate(header_row)
                    if header is not None and str(header) not in drop_columns
                ]
                dest_headers = sheet_headers.setdefault(name, [])
                if (
                    source_file is not None
                    or any(header == "Source_File" for _, header in indexed_headers)
                ) and "Source_File" not in dest_headers:
                    dest_headers.insert(0, "Source_File")
                for _, header in indexed_headers:
                    if header not in dest_headers:
                        dest_headers.append(header)

                dest_rows = sheet_rows.setdefault(name, [])
                for row in rows:
                    record = {
                        header: row[index] if index < len(row) else None
                        for index, header in indexed_headers
                    }
                    existing_source = record.get("Source_File")
                    if (
                        source_file is not None
                        and (existing_source is None or str(existing_source).strip() == "")
                    ):
                        record["Source_File"] = source_file
                    dest_rows.append(record)
        finally:
            wb.close()

    for name, headers in sheet_headers.items():
        if "Source_File" in headers and headers[0] != "Source_File":
            headers = ["Source_File"] + [
                header for header in headers if header != "Source_File"
            ]
        dest = merged.create_sheet(name)
        _append_text_safe(dest, headers)
        for record in sheet_rows.get(name, []):
            _append_text_safe(dest, [record.get(header) for header in headers])

    # Apply styling to every sheet (including Processing Info).
    for sheet in merged.worksheets:
        _apply_styles(sheet)
        if sheet.title == "Processing Info":
            _style_processing_info(sheet)

    buf = io.BytesIO()
    merged.save(buf)
    # keep_vba leaves the template's zip handle open; release it.
    if merged.vba_archive is not None:
        merged.vba_archive.close()
    return buf.getvalue()




# -------------------------------------------
# Internal Helper Functions
# -------------------------------------------

def _append_text_safe(sheet: Worksheet, values: list[object]) -> None:
    """Append a row, keeping '='-prefixed strings as text.

    openpyxl coerces any string starting with '=' into a live formula;
    per-job workbook cells are server-provided, so force them back to
    inline strings to block formula injection into the merged workbook.
    """
    sheet.append(values)
    for cell in sheet[sheet.max_row]:
        if isinstance(cell.value, str) and cell.value.startswith("="):
            cell.data_type = "s"


def _apply_styles(sheet: Worksheet) -> None:
    """Apply backend-matching styles in place.

    Header row → blue fill, white bold, borders, center+wrap.
    Data rows  → alt-row fill on even rows, thin borders, alignment by column name.
    Flagged   → cells where `require_human_check` is true get yellow fill.
    """
    col_count = sheet.max_column
    row_count = sheet.max_row
    if row_count == 0 or col_count == 0:
        return

    # Read header names for column-specific styling.
    col_names: list[str] = []
    for c in range(1, col_count + 1):
        v = sheet.cell(row=1, column=c).value
        col_names.append("" if v is None else str(v))

    # Header row.
    for c in range(1, col_count + 1):
        cell = sheet.cell(row=1, column=c)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.border = _HEADER_BORDER
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True,
        )

    max_header_len = max((len(n) for n in col_names), default=10)
    sheet.row_dimensions[1].height = max(30, 25 + (max_header_len // 20) * 15)

    # Data rows.
    for r in range(2, row_count + 1):
        is_even = (r % 2) == 0
        for c in range(1, col_count + 1):
            cell = sheet.cell(row=r, column=c)
            col_name = col_names[c - 1] if c - 1 < len(col_names) else ""
            cell.border = _DATA_BORDER
            if is_even:
                cell.fill = _ALT_ROW_FILL
            if col_name in _WRAP_TEXT_COLS:
                cell.alignment = Alignment(
                    horizontal="left", vertical="center", wrap_text=True,
                )
            elif col_name in _LEFT_ALIGN_COLS:
                cell.alignment = Alignment(horizontal="left", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            if col_name == "require_human_check" and (
                cell.value is True or str(cell.value) == "True"
            ):
                cell.fill = _FLAGGED_FILL
                cell.font = _FLAGGED_FONT

    # Auto-fit column widths (cap at 60).
    for c in range(1, col_count + 1):
        max_len = 0
        for r in range(1, row_count + 1):
            v = sheet.cell(row=r, column=c).value
            if v is not None:
                max_len = max(max_len, len(str(v)))
        sheet.column_dimensions[get_column_letter(c)].width = min(max_len + 2, 60)

    # Freeze header row.
    sheet.freeze_panes = "A2"


def _style_processing_info(sheet: Worksheet) -> None:
    """Red+bold on '— Failed Boreholes —', yellow fill on failed-name rows.

    Mirrors backend `_style_processing_info` in
    `Agentic 06/.../data_processing_utils.py`. Runs after `_apply_styles`
    so the yellow fill overrides the alt-row gray.
    """
    in_failed_section = False
    for r in range(2, sheet.max_row + 1):
        cell_value = str(sheet.cell(row=r, column=1).value or "")
        if cell_value == "— Failed Boreholes —":
            sheet.cell(row=r, column=1).font = _FAILED_HEADER_FONT
            in_failed_section = True
            continue
        if in_failed_section:
            if cell_value.startswith("  ") and cell_value.strip():
                sheet.cell(row=r, column=1).fill = _FAILED_NAME_FILL
                sheet.cell(row=r, column=2).fill = _FAILED_NAME_FILL
            elif cell_value.strip() and not cell_value.startswith("  "):
                in_failed_section = False
