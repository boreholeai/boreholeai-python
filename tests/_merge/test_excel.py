"""Unit tests for Excel data-sheet merge + styling."""

from __future__ import annotations

import io
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

from boreholeai._merge._excel import _apply_styles, merge_excel_files


def _write_book(tmp_path: Path, name: str, sheets: dict[str, list[list]]) -> Path:
    wb = Workbook(); wb.remove(wb.active)
    for sheet_name, rows in sheets.items():
        ws = wb.create_sheet(sheet_name)
        for row in rows:
            ws.append(row)
    p = tmp_path / name
    wb.save(p)
    return p


def test_merge_appends_rows_skipping_header(tmp_path):
    a = _write_book(tmp_path, "a.xlsx", {
        "Material": [["Hole_ID", "from"], ["BH1", "0"]],
    })
    b = _write_book(tmp_path, "b.xlsx", {
        "Material": [["Hole_ID", "from"], ["BH2", "0"], ["BH3", "0"]],
    })

    out = io.BytesIO(merge_excel_files([a, b]))
    wb = load_workbook(out, data_only=True)
    rows = list(wb["Material"].iter_rows(values_only=True))
    assert rows[0] == ("Hole_ID", "from")  # header copied once
    data = sorted(r[0] for r in rows[1:])
    assert data == ["BH1", "BH2", "BH3"]


def test_drop_columns_removes_page_from_every_sheet(tmp_path):
    a = _write_book(tmp_path, "a.xlsx", {
        "SPT": [["depth_top", "SPT_N", "page"], [1.0, 10, 1]],
    })
    b = _write_book(tmp_path, "b.xlsx", {
        "SPT": [["depth_top", "SPT_N", "page"], [2.0, 20, 3]],
    })

    out = io.BytesIO(merge_excel_files([a, b], drop_columns=frozenset({"page"})))
    wb = load_workbook(out, data_only=True)
    rows = list(wb["SPT"].iter_rows(values_only=True))
    assert rows[0] == ("depth_top", "SPT_N")
    assert sorted(r for r in rows[1:]) == [(1.0, 10), (2.0, 20)]


def test_drop_columns_defaults_to_keeping_everything(tmp_path):
    a = _write_book(tmp_path, "a.xlsx", {
        "SPT": [["depth_top", "page"], [1.0, 1]],
    })

    out = io.BytesIO(merge_excel_files([a]))
    wb = load_workbook(out, data_only=True)
    assert list(wb["SPT"].iter_rows(values_only=True))[0] == ("depth_top", "page")


def test_mixed_schema_merge_is_header_aware_and_backfills_source_file(tmp_path):
    legacy = _write_book(tmp_path, "legacy.xlsx", {
        "Material": [["Hole_ID", "from_mbgl"], ["BH1", 0.0]],
    })
    current = _write_book(tmp_path, "current.xlsx", {
        "Material": [
            ["Hole_ID", "Source_File", "to_mbgl", "from_mbgl"],
            ["BH2", "backend-file", 2.0, 1.0],
        ],
    })

    out = io.BytesIO(merge_excel_files(
        [legacy, current],
        source_files={legacy: "legacy-file", current: "mapping-file"},
    ))
    wb = load_workbook(out, data_only=True)
    rows = list(wb["Material"].iter_rows(values_only=True))

    assert rows[0] == ("Source_File", "Hole_ID", "from_mbgl", "to_mbgl")
    assert rows[1] == ("legacy-file", "BH1", 0.0, None)
    assert rows[2] == ("backend-file", "BH2", 1.0, 2.0)


def test_processing_info_aggregated(tmp_path):
    a = _write_book(tmp_path, "a.xlsx", {
        "Processing Info": [["Metric", "Value"], ["Total Boreholes", "1"]],
        "Material": [["Hole_ID"], ["BH1"]],
    })
    b = _write_book(tmp_path, "b.xlsx", {
        "Processing Info": [["Metric", "Value"], ["Total Boreholes", "2"]],
        "Material": [["Hole_ID"], ["BH2"]],
    })

    out = io.BytesIO(merge_excel_files([a, b]))
    wb = load_workbook(out, data_only=True)
    pi = {r[0]: r[1] for r in wb["Processing Info"].iter_rows(values_only=True) if r[0]}
    assert pi["Total Boreholes"] == "3"


def test_unique_sheet_only_in_one_input_passes_through(tmp_path):
    a = _write_book(tmp_path, "a.xlsx", {
        "Material": [["Hole_ID"], ["BH1"]],
        "OnlyInA": [["X"], ["a"]],
    })
    b = _write_book(tmp_path, "b.xlsx", {
        "Material": [["Hole_ID"], ["BH2"]],
    })

    out = io.BytesIO(merge_excel_files([a, b]))
    wb = load_workbook(out, data_only=True)
    assert "OnlyInA" in wb.sheetnames


def test_apply_styles_header_and_alt_rows():
    wb = Workbook(); ws = wb.active
    ws.append(["Hole_ID", "material"])
    ws.append(["BH1", "CLAY"])
    ws.append(["BH2", "SAND"])  # row 3, odd
    ws.append(["BH3", "GRAVEL"])  # row 4, even — should get alt fill

    _apply_styles(ws)

    # Header
    assert ws.cell(1, 1).font.bold is True
    assert ws.freeze_panes == "A2"

    # Alt-row fill on row 4 (even), not row 3 (odd)
    row4_fill = ws.cell(4, 1).fill.fgColor.value
    row3_fill = ws.cell(3, 1).fill.fgColor.value
    assert row4_fill == "FFF8F9FA"
    # Row 3 has no fgColor set explicitly → openpyxl returns "00000000"
    assert row3_fill in (None, "00000000")


def test_apply_styles_flags_require_human_check():
    wb = Workbook(); ws = wb.active
    ws.append(["Hole_ID", "require_human_check"])
    ws.append(["BH1", True])

    _apply_styles(ws)

    flagged = ws.cell(2, 2)
    assert flagged.fill.fgColor.value == "FFFFEB9C"


def test_empty_input_raises():
    with pytest.raises(ValueError):
        merge_excel_files([])
