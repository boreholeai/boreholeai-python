"""Macro-template merge: VBA project + button parts must survive the merge.

The real template is authored once in Excel (see templates/README.md). These
tests build a structural stand-in — a valid workbook whose zip is augmented
with a stub vbaProject.bin, the button's VML drawing, and its ctrlProp part
— and assert the merge carries all of it into the output so the authored
template's macro and button survive the openpyxl round-trip.
"""

from __future__ import annotations

import re
import shutil
import zipfile
from pathlib import Path

from openpyxl import Workbook, load_workbook

import boreholeai._merge as merge_mod
from boreholeai._merge import merge_results
from boreholeai._merge._excel import merge_excel_files

_STUB_VBA = b"stub-vba-project-bytes"

_BUTTON_VML = """<xml xmlns:v="urn:schemas-microsoft-com:vml"
 xmlns:o="urn:schemas-microsoft-com:office:office"
 xmlns:x="urn:schemas-microsoft-com:office:excel">
 <v:shape id="_x0000_s1025" type="#_x0000_t201" o:button="t"
  style="position:absolute;margin-left:230pt;margin-top:40pt;width:150pt;height:40pt">
  <x:ClientData ObjectType="Button">
   <x:Anchor>3, 5, 2, 5, 4, 40, 5, 10</x:Anchor>
   <x:FmlaMacro>[0]!RegenerateDerivedTabs</x:FmlaMacro>
  </x:ClientData>
 </v:shape>
</xml>"""

_CTRL_PROP = (
    '<formControlPr xmlns="http://schemas.microsoft.com/office/'
    'spreadsheetml/2009/9/main" objectType="Button" lockText="1"/>'
)

_SHEET_RELS = """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
  <Relationship Id="rId2" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/vmlDrawing" Target="../drawings/vmlDrawing1.vml"/>
  <Relationship Id="rId3" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/ctrlProp" Target="../ctrlProps/ctrlProp1.xml"/>
</Relationships>"""


def _make_stub_template(tmp_path: Path) -> Path:
    """Build a keep_vba-loadable stand-in for ground_profile_template.xlsm."""
    base = tmp_path / "template_base.xlsx"
    wb = Workbook()
    wb.active.title = "Processing Info"
    wb.save(base)

    out = tmp_path / "ground_profile_template.xlsm"
    with zipfile.ZipFile(base) as zin, zipfile.ZipFile(out, "w") as zout:
        for item in zin.namelist():
            data = zin.read(item)
            if item == "[Content_Types].xml":
                text = data.decode()
                text = text.replace(
                    "application/vnd.openxmlformats-officedocument"
                    ".spreadsheetml.sheet.main+xml",
                    "application/vnd.ms-excel.sheet.macroEnabled.main+xml",
                )
                text = text.replace(
                    "</Types>",
                    '<Default Extension="bin" ContentType="application/vnd.ms-office.vbaProject"/>'
                    '<Default Extension="vml" ContentType="application/vnd.openxmlformats-officedocument.vmlDrawing"/>'
                    '<Override PartName="/xl/ctrlProps/ctrlProp1.xml" ContentType="application/vnd.ms-excel.controlproperties+xml"/>'
                    "</Types>",
                )
                data = text.encode()
            elif item == "xl/_rels/workbook.xml.rels":
                data = data.decode().replace(
                    "</Relationships>",
                    '<Relationship Id="rIdVba" Type="http://schemas.microsoft.com/office/2006/relationships/vbaProject" Target="vbaProject.bin"/>'
                    "</Relationships>",
                ).encode()
            elif item == "xl/worksheets/sheet1.xml":
                # Excel declares xmlns:r on the worksheet root; the openpyxl-
                # written base doesn't, so declare it on the element itself.
                data = data.decode().replace(
                    "</worksheet>",
                    '<legacyDrawing xmlns:r="http://schemas.openxmlformats.org'
                    '/officeDocument/2006/relationships" r:id="rId2"/>'
                    "</worksheet>",
                ).encode()
            zout.writestr(item, data)
        zout.writestr("xl/vbaProject.bin", _STUB_VBA)
        zout.writestr("xl/drawings/vmlDrawing1.vml", _BUTTON_VML)
        zout.writestr("xl/ctrlProps/ctrlProp1.xml", _CTRL_PROP)
        zout.writestr("xl/worksheets/_rels/sheet1.xml.rels", _SHEET_RELS)
    return out


def _write_ground_profile(tmp_path: Path, name: str) -> Path:
    wb = Workbook()
    pi = wb.active
    pi.title = "Processing Info"
    pi.append(["Metric", "Value"])
    pi.append(["Final Material Records", "2"])
    mat = wb.create_sheet("Material")
    mat.append(["Hole_ID", "from_mbgl", "to_mbgl", "geology_type"])
    mat.append(["BH1", 0.0, 1.0, "FILL"])
    mat.append(["BH1", 1.0, 2.0, "ALLUVIUM"])
    geo = wb.create_sheet("Geology")
    geo.append(["Hole_ID", "from_mbgl", "to_mbgl", "geology_type"])
    geo.append(["BH1", 0.0, 1.0, "FILL"])
    path = tmp_path / name
    wb.save(path)
    return path


def test_macro_template_parts_survive_merge(tmp_path):
    template = _make_stub_template(tmp_path)
    p1 = _write_ground_profile(tmp_path, "job1.xlsx")
    p2 = _write_ground_profile(tmp_path, "job2.xlsx")

    blob = merge_excel_files(
        [p1, p2],
        source_files={p1: "fileA", p2: "fileB"},
        macro_template=template,
    )
    out = tmp_path / "merged.xlsm"
    out.write_bytes(blob)

    with zipfile.ZipFile(out) as z:
        names = set(z.namelist())
        assert "xl/vbaProject.bin" in names
        assert z.read("xl/vbaProject.bin") == _STUB_VBA
        assert "xl/drawings/vmlDrawing1.vml" in names
        assert b"RegenerateDerivedTabs" in z.read("xl/drawings/vmlDrawing1.vml")
        assert "xl/ctrlProps/ctrlProp1.xml" in names
        # The button's sheet must still reference its VML part.
        sheet_xmls = [
            z.read(n) for n in names
            if re.fullmatch(r"xl/worksheets/sheet\d+\.xml", n)
        ]
        assert any(b"<legacyDrawing" in x for x in sheet_xmls)
        # Workbook part must stay macro-enabled.
        assert (
            b"application/vnd.ms-excel.sheet.macroEnabled.main+xml"
            in z.read("[Content_Types].xml")
        )

    # Data must have landed in the template's own Processing Info sheet,
    # not a duplicate, and the data sheets must merge as usual.
    wb = load_workbook(out, keep_vba=True)
    assert wb.sheetnames.count("Processing Info") == 1
    pi_rows = list(wb["Processing Info"].iter_rows(values_only=True))
    assert any(row[0] == "Final Material Records" for row in pi_rows)
    material = list(wb["Material"].iter_rows(values_only=True))
    assert material[0][0] == "Source_File"
    assert len(material) == 5  # header + 2 rows per job
    wb.vba_archive.close()


def test_merge_results_macro_button_writes_xlsm(tmp_path, monkeypatch):
    template = _make_stub_template(tmp_path)
    monkeypatch.setattr(merge_mod, "_MACRO_TEMPLATE_PATH", template)

    job_dirs = []
    for i in (1, 2):
        d = tmp_path / f"job{i}"
        d.mkdir()
        _write_ground_profile(d, "Borehole_ground_profile.xlsx")
        job_dirs.append(d)

    out_dir = tmp_path / "out"
    result = merge_results(job_dirs, out_dir, macro_button=True)
    produced = {p.name for p in result.files}
    assert "Borehole_ground_profile_merged.xlsm" in produced
    assert "Borehole_ground_profile_merged.xlsx" not in produced


def test_merge_results_macro_button_single_dir_writes_xlsm(tmp_path, monkeypatch):
    template = _make_stub_template(tmp_path)
    monkeypatch.setattr(merge_mod, "_MACRO_TEMPLATE_PATH", template)

    d = tmp_path / "job1"
    d.mkdir()
    _write_ground_profile(d, "Borehole_ground_profile.xlsx")

    out_dir = tmp_path / "out"
    result = merge_results([d], out_dir, macro_button=True)
    produced = {p.name for p in result.files}
    assert "Borehole_ground_profile.xlsm" in produced
    assert "Borehole_ground_profile.xlsx" not in produced


def test_merge_results_missing_template_falls_back_to_xlsx(tmp_path, monkeypatch):
    monkeypatch.setattr(
        merge_mod, "_MACRO_TEMPLATE_PATH", tmp_path / "missing.xlsm",
    )
    d = tmp_path / "job1"
    d.mkdir()
    _write_ground_profile(d, "Borehole_ground_profile.xlsx")

    result = merge_results([d], tmp_path / "out", macro_button=True)
    produced = {p.name for p in result.files}
    assert "Borehole_ground_profile.xlsx" in produced
    assert not any(name.endswith(".xlsm") for name in produced)
    assert any("macro button unavailable" in w for w in result.warnings)


def test_merge_results_corrupt_template_falls_back_to_xlsx(tmp_path, monkeypatch):
    corrupt = tmp_path / "ground_profile_template.xlsm"
    corrupt.write_bytes(b"this is not a zip archive")
    monkeypatch.setattr(merge_mod, "_MACRO_TEMPLATE_PATH", corrupt)

    job_dirs = []
    for i in (1, 2):
        d = tmp_path / f"job{i}"
        d.mkdir()
        _write_ground_profile(d, "Borehole_ground_profile.xlsx")
        job_dirs.append(d)

    result = merge_results(job_dirs, tmp_path / "out", macro_button=True)
    produced = {p.name for p in result.files}
    assert "Borehole_ground_profile_merged.xlsx" in produced
    assert not any(name.endswith(".xlsm") for name in produced)
    assert any("macro button unavailable" in w for w in result.warnings)
    # The fallback workbook must still be a complete, loadable merge.
    out = tmp_path / "out" / "Borehole_ground_profile_merged.xlsx"
    wb = load_workbook(out)
    assert len(list(wb["Material"].iter_rows(values_only=True))) == 5


def test_macro_button_false_gives_plain_xlsx(tmp_path):
    d = tmp_path / "job1"
    d.mkdir()
    _write_ground_profile(d, "Borehole_ground_profile.xlsx")

    out_dir = tmp_path / "out"
    result = merge_results([d], out_dir, macro_button=False)
    produced = {p.name for p in result.files}
    assert "Borehole_ground_profile.xlsx" in produced
    assert not any(name.endswith(".xlsm") for name in produced)
    assert not any("macro" in w for w in result.warnings)


def test_macro_button_default_on_uses_packaged_template(tmp_path):
    # No monkeypatching: the default must pick up the real authored
    # template that ships inside the package.
    d = tmp_path / "job1"
    d.mkdir()
    _write_ground_profile(d, "Borehole_ground_profile.xlsx")

    out_dir = tmp_path / "out"
    result = merge_results([d], out_dir)
    produced = {p.name for p in result.files}
    assert "Borehole_ground_profile.xlsm" in produced
    assert not any("macro" in w for w in result.warnings)
    with zipfile.ZipFile(out_dir / "Borehole_ground_profile.xlsm") as z:
        assert "xl/vbaProject.bin" in z.namelist()
