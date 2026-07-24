"""Unit tests for Processing Info aggregation."""

from __future__ import annotations

from openpyxl import Workbook

from boreholeai._merge._processing_info import (
    _format_duration,
    _parse_duration,
    _parse_num,
    build_merged_processing_info,
    read_processing_info,
)


def test_parse_num_integers():
    assert _parse_num("5") == 5
    assert _parse_num("123") == 123


def test_parse_num_handles_garbage():
    assert _parse_num(None) == 0
    assert _parse_num("") == 0
    assert _parse_num("N/A") == 0
    assert _parse_num("Total: 5") == 5  # strips letters


def test_parse_duration():
    assert _parse_duration("5 min") == 300.0
    assert _parse_duration("30 sec") == 30.0
    assert _parse_duration("5 min 30 sec") == 330.0
    assert _parse_duration("N/A") == 0.0
    assert _parse_duration(None) == 0.0


def test_format_duration_roundtrip():
    for sec in [0, 30, 60, 330, 600, 3600]:
        if sec == 0:
            assert _format_duration(sec) == "N/A"
            continue
        formatted = _format_duration(sec)
        assert _parse_duration(formatted) == sec


def test_format_duration_under_a_minute():
    assert _format_duration(45) == "45 sec"


def test_format_duration_minutes_and_seconds():
    assert _format_duration(330) == "5 min 30 sec"


def test_read_processing_info_skips_blank_rows():
    wb = Workbook()
    ws = wb.active
    ws.append(["Metric", "Value"])
    ws.append([None, None])  # blank — should be skipped
    ws.append(["Total Boreholes", "3"])
    ws.append(["Total Pages Processed", "10"])

    info = read_processing_info(ws)
    assert info["Total Boreholes"] == "3"
    assert info["Total Pages Processed"] == "10"


def test_build_merged_aggregates_totals():
    info_a = {"Total Boreholes": "3", "Total Pages Processed": "10",
              "Total Processing Time": "5 min", "Material Records": "12"}
    info_b = {"Total Boreholes": "2", "Total Pages Processed": "5",
              "Total Processing Time": "3 min 30 sec", "Material Records": "8"}

    wb = Workbook(); dest = wb.active
    build_merged_processing_info([info_a, info_b], dest)

    rows = {r[0]: r[1] for r in dest.iter_rows(values_only=True) if r[0]}
    assert rows["Total Boreholes"] == "5"
    assert rows["Total Pages Processed"] == "15"
    assert rows["Total Processing Time"] == "8 min 30 sec"
    assert rows["Material Records"] == "20"


def test_build_merged_collects_failed_boreholes():
    """Failed entries are indented (start with two spaces) and listed under their header."""
    info = {"Total Boreholes": "2", "Boreholes Failed": "1",
            "  BH99": ""}  # the failed-entry convention

    wb = Workbook(); dest = wb.active
    build_merged_processing_info([info], dest)

    metrics = [r[0] for r in dest.iter_rows(values_only=True) if r[0]]
    assert "  BH99" in metrics


def test_build_merged_carries_failure_reason():
    """A failed borehole's reason (column B) is preserved in the merged sheet."""
    reason = "Page 3: Could not read the borehole log on this page"
    info = {"Total Boreholes": "2", "Boreholes Failed": "1", "  BH01": reason}

    wb = Workbook(); dest = wb.active
    build_merged_processing_info([info], dest)

    rows = {r[0]: r[1] for r in dest.iter_rows(values_only=True) if r[0]}
    assert rows["  BH01"] == reason


def test_build_merged_combines_duplicate_failure_reasons():
    """Same borehole across two files: reasons are combined with '; '."""
    info_a = {"Boreholes Failed": "1", "  BH01": "Page 1: reason A"}
    info_b = {"Boreholes Failed": "1", "  BH01": "Page 2: reason B"}

    wb = Workbook(); dest = wb.active
    build_merged_processing_info([info_a, info_b], dest)

    rows = {r[0]: r[1] for r in dest.iter_rows(values_only=True) if r[0]}
    assert rows["  BH01"] == "Page 1: reason A; Page 2: reason B"


def test_build_merged_dedups_identical_failure_reason():
    """Identical reason for the same borehole is not repeated."""
    reason = "Page 1: reason A"
    info_a = {"  BH01": reason}
    info_b = {"  BH01": reason}

    wb = Workbook(); dest = wb.active
    build_merged_processing_info([info_a, info_b], dest)

    rows = {r[0]: r[1] for r in dest.iter_rows(values_only=True) if r[0]}
    assert rows["  BH01"] == reason


def test_build_merged_failed_borehole_blank_reason_compat():
    """Older workbooks store a blank reason — the merged cell stays blank."""
    info = {"Boreholes Failed": "1", "  BH99": ""}

    wb = Workbook(); dest = wb.active
    build_merged_processing_info([info], dest)

    rows = {r[0]: r[1] for r in dest.iter_rows(values_only=True) if r[0]}
    assert rows["  BH99"] in ("", None)


def test_build_merged_handles_zero_pages():
    """Avg Time per Page must be 'N/A' when pages == 0 (no division by zero)."""
    info = {"Total Boreholes": "0", "Total Pages Processed": "0",
            "Total Processing Time": "N/A"}

    wb = Workbook(); dest = wb.active
    build_merged_processing_info([info], dest)

    rows = {r[0]: r[1] for r in dest.iter_rows(values_only=True) if r[0]}
    assert rows["Average Time per Page"] == "N/A"


def _make_json_doc(total, failed, name, reason):
    return {
        "ground_profile": {
            "processing_info": [
                {"Metric": "— Processing Summary —", "Value": ""},
                {"Metric": "Total Boreholes", "Value": str(total)},
                {"Metric": "Boreholes Digitalised", "Value": str(total - failed)},
                {"Metric": "Boreholes Failed", "Value": str(failed)},
                {"Metric": "— Failed Boreholes —", "Value": ""},
                {"Metric": f"  {name}", "Value": reason},
                {"Metric": "— Record Counts —", "Value": ""},
                {"Metric": "Material Records", "Value": "5"},
            ],
            "material": [{"Hole_ID": name, "x": 1}],
        }
    }


def test_merge_json_aggregates_processing_info_like_excel(tmp_path):
    """JSON merge aggregates processing_info (dedup summary, combine reasons,
    sum counts) instead of naively concatenating, and drops the worksheet
    header. Record sections are still concatenated."""
    import json

    from boreholeai._merge._json import merge_json_files

    a = tmp_path / "a.json"
    a.write_text(json.dumps(_make_json_doc(3, 1, "BH01", "Page 1: reason A")))
    b = tmp_path / "b.json"
    b.write_text(json.dumps(_make_json_doc(2, 1, "BH02", "Page 2: reason B")))

    out = json.loads(merge_json_files([a, b]))
    pi = out["ground_profile"]["processing_info"]

    # No worksheet header leaked into the JSON records.
    assert not any(r["Metric"] == "Metric" and r["Value"] == "Value" for r in pi)

    info = {r["Metric"].strip(): r["Value"] for r in pi if r["Metric"].strip()}
    # Summary aggregated once, not duplicated.
    assert sum(1 for r in pi if r["Metric"] == "Total Boreholes") == 1
    assert info["Total Boreholes"] == "5"
    assert info["Boreholes Failed"] == "2"
    # Both failure reasons carried.
    assert info["BH01"] == "Page 1: reason A"
    assert info["BH02"] == "Page 2: reason B"
    # Record counts summed.
    assert info["Material Records"] == "10"
    # Record sections still concatenated.
    assert len(out["ground_profile"]["material"]) == 2


def test_merge_json_drops_page_from_test_data_only(tmp_path):
    """The per-source-file "page" field is dropped from merged test_data
    records but kept on ground_profile records."""
    import json

    from boreholeai._merge._json import merge_json_files

    doc = {
        "ground_profile": {"material": [{"Hole_ID": "BH01", "page": 2}]},
        "test_data": {"SPT": [{"depth_top": 1.0, "SPT_N": 10, "page": 2}]},
    }
    a = tmp_path / "a.json"
    a.write_text(json.dumps(doc))

    out = json.loads(merge_json_files([a]))
    assert out["test_data"]["SPT"] == [{"depth_top": 1.0, "SPT_N": 10}]
    assert out["ground_profile"]["material"] == [{"Hole_ID": "BH01", "page": 2}]


def test_merge_json_backfills_source_file_but_preserves_backend_value(tmp_path):
    import json

    from boreholeai._merge._json import merge_json_files

    legacy = tmp_path / "legacy.json"
    legacy.write_text(json.dumps({
        "ground_profile": {
            "processing_info": [{"Metric": "Total Boreholes", "Value": "1"}],
            "material": [{"Hole_ID": "BH01"}],
        },
    }))
    current = tmp_path / "current.json"
    current.write_text(json.dumps({
        "ground_profile": {
            "material": [{
                "Hole_ID": "BH02",
                "Source_File": "backend-file",
            }],
        },
    }))

    out = json.loads(merge_json_files(
        [legacy, current],
        source_files={legacy: "legacy-file", current: "mapping-file"},
    ))
    rows = out["ground_profile"]["material"]

    assert next(iter(rows[0])) == "Source_File"
    assert rows[0]["Source_File"] == "legacy-file"
    assert next(iter(rows[1])) == "Source_File"
    assert rows[1]["Source_File"] == "backend-file"
    assert "Source_File" not in out["ground_profile"]["processing_info"][0]
