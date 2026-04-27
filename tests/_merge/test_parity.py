"""Parity test: Python merge output vs frontend TS reference output.

Inputs come from real per-job downloads; reference comes from the
frontend's "Download Selected" run (post CSV-parser fix). We verify
semantic equality (group/sheet sets, headers, data row multisets) —
not byte equality. Styling and the "Report Generated" date are
explicitly excluded.
"""

from __future__ import annotations

from pathlib import Path

import openpyxl
import pytest

from boreholeai._merge import merge_results
from boreholeai._merge._ags import _parse_ags

FIXTURES = Path(__file__).parent / "fixtures"
PER_JOB = FIXTURES / "per_job"
REFERENCE = FIXTURES / "reference"


@pytest.fixture(scope="module")
def merged_outputs(tmp_path_factory) -> Path:
    """Run merge_results on the fixture per-job dirs once for the module."""
    out = tmp_path_factory.mktemp("parity_out")
    job_dirs = sorted(d for d in PER_JOB.iterdir() if d.is_dir())
    assert len(job_dirs) >= 2, "need at least 2 per-job fixtures"
    merge_results(job_dirs, out)
    return out


# --- AGS parity ---

def _ags_groups(text: str) -> dict[str, dict]:
    return {
        g.name: {"headings": tuple(g.headings), "rows": {tuple(r) for r in g.data}}
        for g in _parse_ags(text)
    }


def test_ags_same_group_set(merged_outputs):
    py = _ags_groups((merged_outputs / "Borehole_ags4_merged.ags").read_text())
    ts = _ags_groups((REFERENCE / "Borehole_ags4_merged.ags").read_text())
    assert set(py) == set(ts)


def test_ags_headings_match_per_group(merged_outputs):
    py = _ags_groups((merged_outputs / "Borehole_ags4_merged.ags").read_text())
    ts = _ags_groups((REFERENCE / "Borehole_ags4_merged.ags").read_text())
    for name in py:
        assert py[name]["headings"] == ts[name]["headings"], f"headings differ in {name}"


def test_ags_data_rows_match_per_group(merged_outputs):
    py = _ags_groups((merged_outputs / "Borehole_ags4_merged.ags").read_text())
    ts = _ags_groups((REFERENCE / "Borehole_ags4_merged.ags").read_text())
    for name in py:
        diff_only_py = py[name]["rows"] - ts[name]["rows"]
        diff_only_ts = ts[name]["rows"] - py[name]["rows"]
        assert not diff_only_py and not diff_only_ts, (
            f"{name}: only-Py={list(diff_only_py)[:1]}  only-TS={list(diff_only_ts)[:1]}"
        )


# --- Excel parity ---

_REPORT_DATE_METRIC = "Report Generated"


def _sheet_rows(ws) -> set[tuple]:
    out = set()
    for r in ws.iter_rows(values_only=True):
        if all(v is None or (isinstance(v, str) and not v.strip()) for v in r):
            continue
        # Drop the date row — it's expected to drift between runs
        if r and r[0] == _REPORT_DATE_METRIC:
            continue
        out.add(tuple("" if v is None else str(v).strip() for v in r))
    return out


@pytest.mark.parametrize("filename", [
    "Borehole_ground_profile_merged.xlsx",
    "Borehole_test_data_merged.xlsx",
])
def test_excel_sheet_names_match(merged_outputs, filename):
    py = openpyxl.load_workbook(merged_outputs / filename, data_only=True)
    ts = openpyxl.load_workbook(REFERENCE / filename, data_only=True)
    assert set(py.sheetnames) == set(ts.sheetnames)


@pytest.mark.parametrize("filename", [
    "Borehole_ground_profile_merged.xlsx",
    "Borehole_test_data_merged.xlsx",
])
def test_excel_data_rows_match(merged_outputs, filename):
    py = openpyxl.load_workbook(merged_outputs / filename, data_only=True)
    ts = openpyxl.load_workbook(REFERENCE / filename, data_only=True)

    for sn in py.sheetnames:
        if sn not in ts.sheetnames:
            continue
        py_rows = _sheet_rows(py[sn])
        ts_rows = _sheet_rows(ts[sn])
        only_py = py_rows - ts_rows
        only_ts = ts_rows - py_rows
        assert not only_py and not only_ts, (
            f"{filename}/{sn} differs:\n  only-Py: {list(only_py)[:1]}\n  only-TS: {list(only_ts)[:1]}"
        )
